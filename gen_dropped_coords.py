import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def dms_to_dd(s):
    m = re.search(r"(\d{1,3})\D+(\d{1,2})'([\d.]+)\"?\s*([NSEW])", str(s), re.I)
    if not m:
        return None
    d, mi, sec, hem = int(m.group(1)), int(m.group(2)), float(m.group(3)), m.group(4).upper()
    dd = d + mi/60 + sec/3600
    return -dd if hem in ('S', 'W') else dd

def clean_coord(v):
    if v is None:
        return None
    s = str(v).strip()
    if re.search(r"\d+\D+\d+'\d", s) or re.search(r'\d[^\d]*[NSEW]\s*$', s, re.I):
        dd = dms_to_dd(s)
        if dd is not None:
            return dd
    cleaned = re.sub(r'[^0-9.\-]', '', s)
    parts = cleaned.split('.')
    if len(parts) > 2:
        cleaned = parts[0] + '.' + parts[1]
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None

def drop_reason(lat_raw, lon_raw):
    """Determine the specific reason a record could not be recovered."""
    lat_s = str(lat_raw).strip() if lat_raw is not None else ''
    lon_s = str(lon_raw).strip() if lon_raw is not None else ''

    # Missing entirely
    if not lat_s or not lon_s:
        return 'Missing coordinate value'

    # DMS format that still failed after robust parse attempt
    if re.search(r"\d+\D+\d+'\d", lat_s) or re.search(r"\d+\D+\d+'\d", lon_s):
        return 'DMS format — could not extract valid Bangladesh coordinates'

    lat = clean_coord(lat_raw)
    lon = clean_coord(lon_raw)

    if lat is None or lon is None:
        return 'Cannot extract numeric value from coordinate string'

    # Identical values
    try:
        raw_lat = float(str(lat_raw).strip())
        raw_lon = float(str(lon_raw).strip())
        if abs(raw_lat - raw_lon) < 0.0001:
            return 'Identical lat and lon values — data entry error'
    except (ValueError, TypeError):
        pass

    # Negative
    if lat < 0 or lon < 0:
        return 'Negative coordinate — Southern/Western hemisphere value'

    # Implausibly large (missing decimal point)
    if lat > 1000000 or lon > 1000000:
        return 'Implausibly large value — likely missing decimal point'

    # Both out of BD but not swappable
    in_bd = lambda la, lo: (20 <= la <= 27 and 88 <= lo <= 93)
    if not in_bd(lat, lon) and not in_bd(lon, lat):
        return f'Out of Bangladesh bounds after all fixes (lat={lat:.6f}, lon={lon:.6f})'

    return 'Unknown — slipped through all recovery paths'


in_bd = lambda la, lo: (20 <= la <= 27 and 88 <= lo <= 93)

# ── Read source ───────────────────────────────────────────────────
wb_src = openpyxl.load_workbook('MapData/ISP pop list.xlsx', read_only=True)
ws_src = wb_src['Sheet2']
rows = list(ws_src.iter_rows(values_only=True))

dropped = []
for r in rows[1:]:
    sl, name, isp_type, license_no, lat_raw, lon_raw, nttn = r

    # Skip records that were successfully used
    try:
        lat, lon = float(lat_raw), float(lon_raw)
        if in_bd(lat, lon):
            continue
    except (TypeError, ValueError):
        pass

    lat = clean_coord(lat_raw)
    lon = clean_coord(lon_raw)

    if lat is not None and lon is not None:
        if in_bd(lat, lon) or in_bd(lon, lat):
            continue  # successfully fixed

    reason = drop_reason(lat_raw, lon_raw)
    dropped.append((sl, name, isp_type, license_no, lat_raw, lon_raw, nttn, reason))

# ── Build output workbook ─────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = 'Dropped Records'

# Styles
HEADER_FILL  = PatternFill('solid', fgColor='1F3864')
REASON_FILL  = PatternFill('solid', fgColor='FFF3CD')
ALT_FILL     = PatternFill('solid', fgColor='F8FAFC')
HDR_FONT     = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BODY_FONT    = Font(name='Calibri', size=10)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=False)
WRAP         = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin         = Side(style='thin', color='E2E8F0')
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['#', 'ISP Name', 'License Type', 'License Number', 'Raw Lat', 'Raw Lon', 'NTTN Provider', 'Drop Reason']
col_widths = [5, 36, 16, 40, 22, 22, 28, 52]

# Header row
ws.row_dimensions[1].height = 22
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = HDR_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w

# Data rows
for i, row in enumerate(dropped, 2):
    fill = REASON_FILL if i % 2 == 0 else ALT_FILL
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font      = BODY_FONT
        cell.fill      = fill
        cell.border    = BORDER
        cell.alignment = CENTER if col == 1 else WRAP
    ws.row_dimensions[i].height = 18

# Freeze header
ws.freeze_panes = 'A2'

# Summary sheet
ws2 = wb_out.create_sheet('Summary')
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 12

summary_hdr_fill = PatternFill('solid', fgColor='1F3864')
for col in range(1, 3):
    c = ws2.cell(row=1, column=col)
    c.font = HDR_FONT; c.fill = summary_hdr_fill; c.alignment = CENTER; c.border = BORDER

ws2.cell(row=1, column=1, value='Drop Reason')
ws2.cell(row=1, column=2, value='Count')

from collections import Counter
reason_counts = Counter(r[7] for r in dropped)
for i, (reason, count) in enumerate(sorted(reason_counts.items(), key=lambda x: -x[1]), 2):
    ws2.cell(row=i, column=1, value=reason).border = BORDER
    ws2.cell(row=i, column=2, value=count).border  = BORDER
    ws2.cell(row=i, column=1).font = BODY_FONT
    ws2.cell(row=i, column=2).font = BODY_FONT
    ws2.cell(row=i, column=2).alignment = CENTER

ws2.cell(row=len(reason_counts)+2, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=len(reason_counts)+2, column=2, value=len(dropped)).font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=len(reason_counts)+2, column=2).alignment = CENTER

out_path = '../Shakib/ISP-POP-Dropped-Coordinates-v2.xlsx'
wb_out.save(out_path)
print(f'Saved {len(dropped)} dropped records to {out_path}')
for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
    print(f'  {count:3d}  {reason}')
