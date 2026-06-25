import openpyxl, re, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

def dms_to_dd(s):
    m = re.search(r"(\d{1,3})\D+(\d{1,2})'([\d.]+)\"?\s*([NSEW])", str(s), re.I)
    if not m: return None
    d, mi, sec, hem = int(m.group(1)), int(m.group(2)), float(m.group(3)), m.group(4).upper()
    dd = d + mi/60 + sec/3600
    return -dd if hem in ('S', 'W') else dd

def clean_coord(v):
    if v is None: return None
    s = str(v).strip()
    if re.search(r"\d+\D+\d+'\d", s) or re.search(r'\d[^\d]*[NSEW]\s*$', s, re.I):
        dd = dms_to_dd(s)
        if dd is not None: return dd
    cleaned = re.sub(r'[^0-9.\-]', '', s)
    parts = cleaned.split('.')
    if len(parts) > 2: cleaned = parts[0] + '.' + parts[1]
    try: return float(cleaned)
    except: return None

def has_double_dot(v):
    return bool(re.search(r'\d\.\.\d', str(v)))

in_bd = lambda la, lo: (20 <= la <= 27 and 88 <= lo <= 93)

# ── Read v2 dropped list ──────────────────────────────────────────
wb = openpyxl.load_workbook('../Shakib/ISP-POP-Dropped-Coordinates-v2.xlsx', read_only=True)
ws = wb['Dropped Records']
rows = list(ws.iter_rows(values_only=True))

skipped_data_error = []
new_features = []
still_dropped = []

for r in rows[1:]:
    sl, name, isp_type, license_no, lat_raw, lon_raw, nttn, reason = r

    # Flag double-dot data entry error — skip, report to user
    if has_double_dot(lat_raw) or has_double_dot(lon_raw):
        skipped_data_error.append(r)
        still_dropped.append((sl, name, isp_type, license_no, lat_raw, lon_raw, nttn,
                               f'Data entry error — double dot in coordinate (e.g. {lat_raw}, {lon_raw})'))
        continue

    lat = clean_coord(lat_raw)
    lon = clean_coord(lon_raw)

    if lat is not None and lon is not None:
        if in_bd(lat, lon):
            pass  # lat/lon already correct order
        elif in_bd(lon, lat):
            lat, lon = lon, lat
        else:
            still_dropped.append((sl, name, isp_type, license_no, lat_raw, lon_raw, nttn,
                                   f'Still out of Bangladesh bounds (lat={lat:.6f}, lon={lon:.6f})'))
            continue
        new_features.append({
            'type': 'Feature',
            'geometry': {'type': 'Point', 'coordinates': [round(lon, 6), round(lat, 6)]},
            'properties': {
                'name': str(name).strip() if name else '',
                'type': str(isp_type).strip() if isp_type else '',
                'nttn': str(nttn).strip() if nttn else '',
            }
        })
    else:
        still_dropped.append((sl, name, isp_type, license_no, lat_raw, lon_raw, nttn,
                               reason if reason else 'Cannot extract numeric value'))

print(f'Data entry errors (skipped): {len(skipped_data_error)}')
for r in skipped_data_error:
    print(f'  SL={r[0]} | {str(r[1])[:40]} | lat={r[4]} lon={r[5]}')

print(f'\nNew features to add: {len(new_features)}')
print(f'Still dropped:       {len(still_dropped)}')

# ── Append new features to existing GeoJSON ───────────────────────
geojson_path = 'public/data/isp-pops.geojson'
with open(geojson_path, 'r', encoding='utf-8') as f:
    geojson = json.load(f)

before = len(geojson['features'])
geojson['features'].extend(new_features)
after = len(geojson['features'])

with open(geojson_path, 'w', encoding='utf-8') as f:
    json.dump(geojson, f, ensure_ascii=False, separators=(',', ':'))

size_kb = os.path.getsize(geojson_path) // 1024
print(f'\nGeoJSON: {before} -> {after} features ({size_kb} KB)')

# ── Write v3 dropped Excel ────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Dropped Records'

HEADER_FILL = PatternFill('solid', fgColor='1F3864')
REASON_FILL = PatternFill('solid', fgColor='FFF3CD')
ALT_FILL    = PatternFill('solid', fgColor='F8FAFC')
HDR_FONT    = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BODY_FONT   = Font(name='Calibri', size=10)
CENTER      = Alignment(horizontal='center', vertical='center')
WRAP        = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin        = Side(style='thin', color='E2E8F0')
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

headers    = ['#', 'ISP Name', 'License Type', 'License Number', 'Raw Lat', 'Raw Lon', 'NTTN Provider', 'Drop Reason']
col_widths = [5, 36, 16, 40, 22, 22, 28, 52]

ws_out.row_dimensions[1].height = 22
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws_out.cell(row=1, column=col, value=h)
    cell.font = HDR_FONT; cell.fill = HEADER_FILL
    cell.alignment = CENTER; cell.border = BORDER
    ws_out.column_dimensions[get_column_letter(col)].width = w

for i, row in enumerate(still_dropped, 2):
    fill = REASON_FILL if i % 2 == 0 else ALT_FILL
    for col, val in enumerate(row, 1):
        cell = ws_out.cell(row=i, column=col, value=val)
        cell.font = BODY_FONT; cell.fill = fill; cell.border = BORDER
        cell.alignment = CENTER if col == 1 else WRAP
    ws_out.row_dimensions[i].height = 18

ws_out.freeze_panes = 'A2'

# Summary sheet
ws2 = wb_out.create_sheet('Summary')
ws2.column_dimensions['A'].width = 52
ws2.column_dimensions['B'].width = 12

shdr = PatternFill('solid', fgColor='1F3864')
for col in range(1, 3):
    c = ws2.cell(row=1, column=col)
    c.font = HDR_FONT; c.fill = shdr; c.alignment = CENTER; c.border = BORDER
ws2.cell(row=1, column=1, value='Drop Reason')
ws2.cell(row=1, column=2, value='Count')

reason_counts = Counter(r[7] for r in still_dropped)
for i, (reason, count) in enumerate(sorted(reason_counts.items(), key=lambda x: -x[1]), 2):
    ws2.cell(row=i, column=1, value=reason).border = BORDER
    ws2.cell(row=i, column=2, value=count).border  = BORDER
    ws2.cell(row=i, column=1).font = BODY_FONT
    ws2.cell(row=i, column=2).font = Font(name='Calibri', size=10)
    ws2.cell(row=i, column=2).alignment = CENTER

total_row = len(reason_counts) + 2
ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=total_row, column=2, value=len(still_dropped)).font = Font(name='Calibri', bold=True, size=10)
ws2.cell(row=total_row, column=2).alignment = CENTER

out_path = '../Shakib/ISP-POP-Dropped-Coordinates-v3.xlsx'
wb_out.save(out_path)
print(f'Saved v3 with {len(still_dropped)} remaining records to {out_path}')
