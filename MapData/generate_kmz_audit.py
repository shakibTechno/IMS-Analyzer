"""
generate_kmz_audit.py
Audit all LineString features in fiber_network_multiple_district.kmz.
Outputs: KMZ_Line_Audit.xlsx  (client-shareable)
"""

import zipfile, re, pathlib, datetime
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
BASE    = pathlib.Path(r"C:\Users\ASUS\Downloads\BTRC-IMS-RFP-Submission-master"
                       r"\BTRC-IMS-RFP-Submission-master\Demo-UI\MapData")
KMZ     = BASE / "Latest Data" / "fiber_network_multiple_district.kmz"
OUT     = BASE / "KMZ_Line_Audit.xlsx"

# ── KML namespace helper ───────────────────────────────────────────────────
NS_STD = "{http://www.opengis.net/kml/2.2}"
NS_GGL = "{http://earth.google.com/kml/2.2}"

def T(tag, root_tag):
    ns = NS_GGL if "earth.google.com" in root_tag else NS_STD
    return ns + tag

# ── Operator detection ─────────────────────────────────────────────────────
OP_PATTERNS = [
    ("GP",    re.compile(r"\bGP\b|Grameenphone", re.I)),
    ("Robi",  re.compile(r"\bRobi\b", re.I)),
    ("BTCL",  re.compile(r"\bBTCL\b", re.I)),
    ("MOTN",  re.compile(r"\bMOTN\b", re.I)),
    ("BL",    re.compile(r"\bBanglalink\b|\bBL\b", re.I)),
    ("BSCCL", re.compile(r"\bBSCCL\b|\bBSCPLC\b", re.I)),
]

def detect_operator(name: str) -> str:
    for op, pat in OP_PATTERNS:
        if pat.search(name):
            return op
    return "Unknown"

def parse_dist(name: str) -> str:
    m = re.search(r"\((\d+\.?\d*)\s*[Kk]m\)", name)
    return m.group(1) if m else ""

def parse_endpoints(name: str):
    """Split 'Origin - Destination (X Km)' → (origin, destination)."""
    clean = re.sub(r"\s*\(\d+\.?\d*\s*[Kk]m\)\s*$", "", name).strip()
    parts = re.split(r"\s+-\s+|\s+–\s+", clean, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return clean, ""

# ── Parse KMZ ─────────────────────────────────────────────────────────────
print("Parsing KMZ …")
with zipfile.ZipFile(KMZ) as z:
    kml_name = next(n for n in z.namelist() if n.endswith(".kml"))
    with z.open(kml_name) as kf:
        tree = ET.parse(kf)

root     = tree.getroot()
root_tag = root.tag

rows = []
for pm in root.iter(T("Placemark", root_tag)):
    name_el = pm.find(T("name", root_tag))
    name    = (name_el.text or "").strip() if name_el is not None else ""

    ls = pm.find(".//" + T("LineString", root_tag))
    if ls is None:
        continue  # skip points

    operator  = detect_operator(name)
    dist      = parse_dist(name)
    origin, dest = parse_endpoints(name)

    rows.append({
        "name":     name,
        "origin":   origin,
        "dest":     dest,
        "dist_km":  dist,
        "operator": operator,
        "status":   "Identified" if operator != "Unknown" else "Label Missing",
    })

print(f"  Total lines : {len(rows)}")

# ── Operator summary ───────────────────────────────────────────────────────
from collections import Counter
op_counts = Counter(r["operator"] for r in rows)
status_counts = Counter(r["status"] for r in rows)

# ── Colour palette ─────────────────────────────────────────────────────────
C = {
    "header_bg":   "1E3A5F",   # dark navy
    "header_fg":   "FFFFFF",
    "sheet2_bg":   "2D6A4F",   # dark green for summary header
    "identified":  "D1FAE5",   # light green
    "missing":     "FEE2E2",   # light red
    "alt_ident":   "ECFDF5",
    "alt_miss":    "FFF1F1",
    "op_GP":       "DBEAFE",   # blue
    "op_Robi":     "FFEDD5",   # orange
    "op_BTCL":     "CFFAFE",   # cyan
    "op_MOTN":     "EDE9FE",   # purple
    "op_BL":       "FEF3C7",   # amber
    "op_BSCCL":    "D1FAE5",   # green
    "op_Unknown":  "FEE2E2",   # red
    "section_bg":  "F1F5F9",
    "total_bg":    "FFF7ED",
}

OP_COLORS = {
    "GP": C["op_GP"], "Robi": C["op_Robi"], "BTCL": C["op_BTCL"],
    "MOTN": C["op_MOTN"], "BL": C["op_BL"], "BSCCL": C["op_BSCCL"],
    "Unknown": C["op_Unknown"],
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(size=10, color="000000"):
    return Font(bold=True, size=size, color=color)

def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Build workbook ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — All Lines
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "All Lines"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# ── Title block ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "KMZ Fiber Line Audit — fiber_network_multiple_district.kmz"
title_cell.font  = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill  = fill(C["header_bg"])
title_cell.alignment = center()
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:H2")
sub = ws["A2"]
sub.value = (f"Generated: {datetime.date.today().strftime('%d %b %Y')}    "
             f"Total lines: {len(rows)}    "
             f"Identified: {status_counts['Identified']}    "
             f"Label Missing: {status_counts['Label Missing']}")
sub.font      = Font(size=10, italic=True, color="374151")
sub.fill      = fill("EFF6FF")
sub.alignment = center()
ws.row_dimensions[2].height = 18

# ── Column headers ──────────────────────────────────────────────────────────
COLS = ["S/N", "Segment Name (from KMZ)", "Origin Node",
        "Destination Node", "Distance (km)", "Detected Operator",
        "Label Status", "Remarks / Client Action Needed"]
COL_WIDTHS = [6, 50, 30, 30, 13, 18, 16, 38]

for ci, (col, w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
    cell = ws.cell(row=3, column=ci, value=col)
    cell.font      = bold(10, "FFFFFF")
    cell.fill      = fill(C["header_bg"])
    cell.alignment = center()
    cell.border    = thin_border()
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.row_dimensions[3].height = 30

# ── Data rows ───────────────────────────────────────────────────────────────
for idx, r in enumerate(rows, start=1):
    row_n  = idx + 3
    op     = r["operator"]
    is_unk = op == "Unknown"

    # Alternating shade within each status group
    if is_unk:
        bg = C["missing"] if idx % 2 == 1 else C["alt_miss"]
    else:
        bg = C["identified"] if idx % 2 == 1 else C["alt_ident"]

    remark = ""
    if is_unk:
        remark = "Please confirm which operator owns / operates this segment"

    values = [idx, r["name"], r["origin"], r["dest"],
              r["dist_km"], op, r["status"], remark]

    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row_n, column=ci, value=val)
        cell.fill   = fill(bg)
        cell.border = thin_border()
        cell.font   = Font(size=9)

        if ci in (1, 5):
            cell.alignment = center()
        else:
            cell.alignment = left()

        # Operator column: colour-code by operator
        if ci == 6:
            cell.fill = fill(OP_COLORS.get(op, bg))
            cell.font = bold(9)
            cell.alignment = center()

        # Status column bold
        if ci == 7:
            cell.font = bold(9, "047857" if not is_unk else "B91C1C")
            cell.alignment = center()

ws.auto_filter.ref = f"A3:{get_column_letter(len(COLS))}{len(rows)+3}"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — Summary
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:D1")
t2 = ws2["A1"]
t2.value     = "Operator Coverage Summary"
t2.font      = Font(bold=True, size=13, color="FFFFFF")
t2.fill      = fill(C["sheet2_bg"])
t2.alignment = center()
ws2.row_dimensions[1].height = 26

ws2.merge_cells("A2:D2")
s2 = ws2["A2"]
s2.value     = (f"Source: fiber_network_multiple_district.kmz    |    "
                f"Total line segments: {len(rows)}")
s2.font      = Font(size=10, italic=True)
s2.fill      = fill("F0FDF4")
s2.alignment = center()

# Header
hdrs = ["Operator", "Segments", "% of Total", "Status"]
widths2 = [20, 14, 14, 30]
for ci, (h, w) in enumerate(zip(hdrs, widths2), start=1):
    cell = ws2.cell(row=3, column=ci, value=h)
    cell.font      = bold(10, "FFFFFF")
    cell.fill      = fill(C["sheet2_bg"])
    cell.alignment = center()
    cell.border    = thin_border()
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.row_dimensions[3].height = 26

# Op rows — sort identified first, Unknown last
op_order = ["GP", "Robi", "BTCL", "MOTN", "BL", "BSCCL", "Unknown"]
for ri, op in enumerate(op_order, start=4):
    cnt = op_counts.get(op, 0)
    pct = f"{cnt / len(rows) * 100:.1f}%"
    status = "Operator identified in name" if op != "Unknown" else "Label MISSING — client verification needed"
    values = [op, cnt, pct, status]
    for ci, val in enumerate(values, start=1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.fill      = fill(OP_COLORS.get(op, "FFFFFF"))
        cell.border    = thin_border()
        cell.alignment = center() if ci != 4 else left()
        cell.font      = bold(10) if ci == 1 else Font(size=10)
        if ci == 4 and op == "Unknown":
            cell.font = Font(size=10, bold=True, color="B91C1C")

# Total row
ri_total = len(op_order) + 4
tot_vals = ["TOTAL", len(rows), "100%", ""]
for ci, val in enumerate(tot_vals, start=1):
    cell = ws2.cell(row=ri_total, column=ci, value=val)
    cell.fill      = fill(C["total_bg"])
    cell.font      = bold(10)
    cell.alignment = center()
    cell.border    = thin_border()

# ── Note block ──────────────────────────────────────────────────────────────
note_row = ri_total + 2
ws2.merge_cells(f"A{note_row}:D{note_row+4}")
note = ws2.cell(row=note_row, column=1)
note.value = (
    "HOW TO USE THIS FILE\n\n"
    "1. Review the 'All Lines' sheet — rows highlighted in RED have a missing operator label.\n"
    "2. For each red row, please confirm: Which company owns or operates this fiber segment?\n"
    "3. Return the file with the 'Detected Operator' column updated for missing rows.\n"
    "4. We will re-import the confirmed data into the IMS map.\n\n"
    "Operators in use: GP = Grameenphone  |  BTCL = Bangladesh Telecommunications Co. Ltd  |  "
    "MOTN = Moon Telecom Network  |  BL = Banglalink  |  BSCCL = Bangladesh Submarine Cable Co. Ltd"
)
note.font      = Font(size=10, color="1E3A5F")
note.fill      = fill("EFF6FF")
note.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True)
note.border    = thin_border()
ws2.row_dimensions[note_row].height   = 20
for r in range(note_row, note_row + 5):
    ws2.row_dimensions[r].height = 20

# ── Save ────────────────────────────────────────────────────────────────────
print(f"\nSaving -> {OUT.name} ...")
wb.save(OUT)
print("Done.")
print(f"\nSummary:")
for op in op_order:
    cnt = op_counts.get(op, 0)
    bar = "#" * int(cnt / len(rows) * 40)
    print(f"  {op:8s} {cnt:5d}  {bar}")
print(f"  {'TOTAL':8s} {len(rows):5d}")
