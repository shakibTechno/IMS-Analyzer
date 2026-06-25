"""
BTRC IMS — Full Data Inventory Excel (4 sheets)
Sheet 1: Received Files       — every physical file received
Sheet 2: Data Types Found     — content analysis per dataset
Sheet 3: Implementation Status — source vs map gap analysis
Sheet 4: Missing for Complete Map — prioritised action list
"""

import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(r"C:\Users\ASUS\Downloads\BTRC-IMS-RFP-Submission-master\BTRC-IMS-RFP-Submission-master\Demo-UI\MapData\BTRC_Full_Data_Inventory.xlsx")

wb = Workbook()

# ── Palette ─────────────────────────────────────────────────────────
P = dict(
    navy="1E3A5F", blue="2563EB", blue_lt="EFF6FF", blue_acc="BFDBFE",
    green="166534", green_lt="F0FDF4", green_acc="BBF7D0",
    amber="92400E", amber_lt="FFFBEB", amber_acc="FDE68A",
    red="991B1B",  red_lt="FEF2F2",  red_acc="FECACA",
    purple="5B21B6", purple_lt="F5F3FF", purple_acc="DDD6FE",
    slate="334155", slate_lt="F8FAFC", slate_acc="E2E8F0",
    white="FFFFFF", gray="94A3B8", border="CBD5E1",
    partial_bg="FFFBEB", partial_fg="92400E",
)

def sd(c=P["border"], s="thin"):  return Side(border_style=s, color=c)
THIN  = Border(left=sd(), right=sd(), top=sd(), bottom=sd())
MED   = Border(left=sd(P["navy"],"medium"), right=sd(P["navy"],"medium"),
               top=sd(P["navy"],"medium"),  bottom=sd(P["navy"],"medium"))

def fill(h):    return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=9, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)

def wcell(c, val, bold=False, sz=9, fg="000000", bg=None,
          align="left", wrap=True, italic=False, border=THIN):
    c.value = val
    c.font  = Font(name="Calibri", bold=bold, size=sz, color=fg, italic=italic)
    if bg: c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top",
                             wrap_text=wrap)
    c.border = border

def title_row(ws, row, text, ncols, bg=P["navy"], fg=P["white"], sz=13, h=26):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=True, sz=sz, fg=fg, bg=bg, align="center", border=MED)
    ws.row_dimensions[row].height = h

def sub_row(ws, row, text, ncols, bg=P["blue"], fg=P["white"], sz=10, h=16):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=False, sz=sz, fg=fg, bg=bg, align="center", italic=True, border=MED)
    ws.row_dimensions[row].height = h

def grp_row(ws, row, text, ncols, bg, fg=P["navy"], h=15):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=True, sz=10, fg=fg, bg=bg, align="left", border=THIN)
    ws.row_dimensions[row].height = h

def hdr_row(ws, row, cols, h=30):
    for ci, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci)
        wcell(c, label, bold=True, sz=10, fg=P["white"], bg=P["blue"],
              align="center", border=THIN)
    ws.row_dimensions[row].height = h

def set_widths(ws, cols):
    for ci, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — Received Files
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Received Files"

COLS1 = [
    ("S/N",              4),
    ("Operator / Source", 20),
    ("Sub-folder",       22),
    ("File Name",        42),
    ("File Format",      11),
    ("Size (KB)",        10),
    ("Records",          12),
    ("File Purpose",     30),
    ("Used on Map?",     13),
]
NC1 = len(COLS1)

title_row(ws1, 1, "BTRC IMS — Received Data Files Inventory", NC1)
sub_row(ws1, 2, "Complete list of all files received from operators and BTRC. FHLFON listed individually (19 files).", NC1)
ws1.row_dimensions[3].height = 5
hdr_row(ws1, 4, COLS1)
set_widths(ws1, COLS1)
ws1.freeze_panes = "A5"

FILES = [
    # (operator, subfolder, filename, format, size_kb, records, purpose, used)
    # ── Grameenphone ──
    ("Grameenphone (GP)", "Latest Data",
     "GP_Tx Information_18-Dec-25.xlsx", "XLSX", 950, "23,864 rows",
     "BTS sites with TX type (Fiber / MW)", "Yes — Grameenphone layer"),
    # ── Robi ──
    ("Robi Axiata", "Latest Data",
     "Robi site list_Nov'25_BTRC_MW & fiber.xlsx", "XLSX", 816, "18,838 rows",
     "BTS sites with MW and Fiber flags", "Yes — Robi layer"),
    # ── Banglalink ──
    ("Banglalink", "Latest Data",
     "BTS Information_Banglalink_Till 15 Dec 25.xlsx", "XLSX", 1550, "15,154 rows",
     "Latest BTS sites with backhaul type and admin data", "Yes — BL BTS Sites layer"),
    ("Banglalink", "Banglalink",
     "3gtower.geojson", "GeoJSON", 12517, "13,208 features",
     "Historical BTS tower sites with 2G/3G/4G generation", "Yes — BL Towers layer"),
    ("Banglalink", "Banglalink",
     "bl-line.geojson", "GeoJSON", 200, "172 features",
     "Historical fiber route lines (32/48/72 core)", "Yes — BL Fiber Lines layer"),
    # ── BTCL ──
    ("BTCL", "Latest Data",
     "GEO SPIRAL DATA STRUCTURE_TEMPLE_FINAL_BTCL.xlsx", "XLSX", 2074, "24,142 rows",
     "Latest network points (CP/HH/HOP/POP/MH) with coordinates", "Yes — BTCL Points layer"),
    ("BTCL", "BTCL",
     "btcl-ponts.geojson", "GeoJSON", 13708, "29,795 features",
     "Historical network nodes (HOP/HH/CP/MH) with full admin data", "Yes — BTCL-OLD Nodes layer"),
    ("BTCL", "BTCL",
     "btcl-nttn-line.geojson", "GeoJSON", 600, "584 features",
     "Historical fiber lines with core count and route details", "Yes — BTCL-OLD Lines layer"),
    ("BTCL", "MapData root",
     "btcl-union-project-location.geojson", "GeoJSON", 285, "966 features",
     "Union-level BTCL fiber project locations", "Yes — Union Projects layer"),
    # ── Fiber@Home FHLFON — 19 files ──
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.shp", "Shapefile (.shp)", 41069, "141,567 geometries",
     "Line geometries — full network", "Partial (see Sheet 3)"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.dbf", "Shapefile (.dbf)", 171429, "141,567 records",
     "Line attribute table (Optr, Line_Type, Path_Along, Core, etc.)", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.shx", "Shapefile (.shx)", 1133, "—",
     "Shape record offset index", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.sbn", "Shapefile (.sbn)", 1386, "—",
     "Spatial index for fast bounding-box queries", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.sbx", "Shapefile (.sbx)", 40, "—",
     "Spatial index header", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.prj", "Shapefile (.prj)", 1, "—",
     "Coordinate system / projection definition", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.shp.xml", "XML Metadata", 56, "—",
     "ISO metadata for the line shapefile", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLine.cpg", "Shapefile (.cpg)", 1, "—",
     "Character encoding file (UTF-8)", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONLineExcel.xlsx", "XLSX", 12852, "141,567 rows",
     "Line data in tabular form — same records as shapefile", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.shp", "Shapefile (.shp)", 3340, "122,150 geometries",
     "Point geometries — full network", "Partial (see Sheet 3)"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.dbf", "Shapefile (.dbf)", 128591, "122,150 records",
     "Point attribute table (point_id, point_type, point_name, address, lat, lon)", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.shx", "Shapefile (.shx)", 954, "—",
     "Shape record offset index", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.sbn", "Shapefile (.sbn)", 1114, "—",
     "Spatial index", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.sbx", "Shapefile (.sbx)", 37, "—",
     "Spatial index header", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.prj", "Shapefile (.prj)", 1, "—",
     "Coordinate system definition", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.shp.xml", "XML Metadata", 1, "—",
     "ISO metadata for the point shapefile", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPoint.cpg", "Shapefile (.cpg)", 1, "—",
     "Character encoding file", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "FHLFONPointExcel.xlsx", "XLSX", 11121, "122,150 rows",
     "Point data tabular — same records as shapefile", "Partial"),
    ("Fiber@Home (FHLFON)", "FHLFON",
     "Copy of FHLFONPointExcel.xlsx", "XLSX", 11121, "122,150 rows",
     "Duplicate of FHLFONPointExcel.xlsx", "No — duplicate"),
    # ── Summit ──
    ("Summit Communications", "Summit/Line_data Excel",
     "Line_data.xlsx", "XLSX", 11310, "102,514 rows",
     "Full fiber line network (Aerial/Burial/PGCB/Bridge)", "Partial (see Sheet 3)"),
    ("Summit Communications", "Summit/Line_data shape",
     "Line_data.shp", "Shapefile (.shp)", 21949, "102,514 geometries",
     "Line geometries — full network", "Partial"),
    ("Summit Communications", "Summit/Line_data shape",
     "Line_data.dbf", "Shapefile (.dbf)", 82091, "102,514 records",
     "Line attributes (operator, line type, path, ducts, cores)", "Partial"),
    ("Summit Communications", "Summit/Line_data shape",
     "Line_data.shx + .sbn + .sbx + .cpg + .prj", "Shapefile (index files)", 858, "—",
     "Shape index, spatial index, encoding, projection", "Partial"),
    ("Summit Communications", "Summit/Line_data shape",
     "Line_data.shp.xml", "XML Metadata", 625, "—",
     "ISO metadata for line shapefile", "Partial"),
    ("Summit Communications", "Summit/Point_data Excel",
     "Point_data.xlsx", "XLSX", 4990, "68,850 rows",
     "Full point network (TJB/HH/BTS/EP/Node/ODB/POP/POC etc.)", "Partial (see Sheet 3)"),
    ("Summit Communications", "Summit/Point_data shape",
     "Point_data.shp", "Shapefile (.shp)", 1882, "68,850 geometries",
     "Point geometries — full network", "Partial"),
    ("Summit Communications", "Summit/Point_data shape",
     "Point_data.dbf", "Shapefile (.dbf)", 40745, "68,850 records",
     "Point attributes (operator, point_type, point_name, lat, lon)", "Partial"),
    ("Summit Communications", "Summit/Point_data shape",
     "Point_data.shx + .sbn + .sbx + .cpg + .prj", "Shapefile (index files)", 584, "—",
     "Index, spatial index, encoding, projection", "Partial"),
    ("Summit Communications", "Railway",
     "Summit.rar", "RAR Archive", 28696, "—",
     "Archived Summit network data via Railway folder", "No — not extracted"),
    # ── Bahon ──
    ("Bahon Limited", "Bahon Limited Shape Files",
     "Bahon Network_System Line.shp", "Shapefile (.shp)", 7179, "~7,763 geometries",
     "Full fiber line network geometry", "Yes — Bahon layer"),
    ("Bahon Limited", "Bahon Limited Shape Files",
     "Bahon Network_System Line.dbf", "Shapefile (.dbf)", 49326, "~7,763 records",
     "Line attributes (cable type OH/UG/WC, admin data)", "Yes"),
    ("Bahon Limited", "Bahon Limited Shape Files",
     "Bahon Network_System Line.shx + .sbn + .sbx + .cpg + .prj + .qmd",
     "Shapefile (index/meta)", 133, "—",
     "Index, spatial index, encoding, projection, QGIS metadata", "Yes"),
    # ── InfoSarkar-3 ──
    ("InfoSarkar-3 (IS3)", "Info Sarker-3 FHL",
     "doc.kml", "KML", 14786, "~3,860 features",
     "Full IS3 fiber network (lines by core count + nodes)", "Yes — InfoSarkar-3 layer"),
    ("InfoSarkar-3 (IS3)", "Info Sarker-3 FHL/files",
     "mysavedplaces_closed.png + mysavedplaces_open.png", "PNG", 2, "—",
     "KML icon assets embedded in the KMZ", "No — icon files only"),
    # ── PGCB ──
    ("PGCB", "MapData root",
     "Power Grid Tranmission(OPGW).kml", "KML", 1128, "324 features",
     "OPGW fiber along power transmission lines (400/230/132kV + UG)", "Yes — PGCB layer"),
    # ── Multi-operator Fiber Network ──
    ("Multi-operator\n(GP, Robi, BTCL, BL, MOTN, BSCCL)", "Latest Data",
     "fiber_network_multiple_district.kmz", "KMZ", 2171, "Lines: 8,163\nPoints: 19,096",
     "Inter-district fiber routes and junction points for 6 operators", "Yes — Fiber Network layer"),
    # ── Bangladesh Railway ──
    ("Bangladesh Railway (BR)", "Latest Data",
     "Geo Spatial Data Structure_Template _Final_railway.xlsx", "XLSX", 103, "353 routes",
     "Railway fiber routes (8–96 core) with station names and core details", "Yes — BR Fiber layer"),
    ("Bangladesh Railway (BR)", "Railway",
     "railway.geojson", "GeoJSON", 52, "2,675 segments",
     "Railway track geometry (base layer)", "Yes — Railline layer"),
    ("Bangladesh Railway (BR)", "Railway/railline_wgs",
     "railline_wgs.shp + .dbf + .shx + .prj + .cst", "Shapefile", 49, "—",
     "Railway track shapefile (source for railway.geojson)", "Indirect"),
    ("Bangladesh Railway (BR)", "Railway",
     "Geo Spatial Data Structure_Template _Final_railway.xlsx", "XLSX", 104, "353 routes",
     "Duplicate of Latest Data file (same content)", "No — duplicate"),
    # ── ISP POP ──
    ("All ISPs (Multi-operator)", "pop all isp",
     "aisp-pop.geojson", "GeoJSON", 2167, "3,930 features",
     "ISP POP locations with ISP name, capacity, admin data (all districts)", "No — not on map yet"),
    ("All ISPs (Multi-operator)", "MapData root",
     "all-isp-pop-info.json", "JSON", 2167, "3,930 records",
     "Same ISP POP data in JSON format", "No — not on map yet"),
]

row = 5
alt = False
prev_op = None
sn = 0
USED_MAP = {
    "Yes":     (P["green_lt"],  P["green"]),
    "Partial": (P["amber_lt"],  P["amber"]),
    "No":      (P["red_lt"],    P["red"]),
    "Indirect":(P["purple_lt"], P["purple"]),
}

for (op, folder, fname, fmt, size_kb, records, purpose, used) in FILES:
    if op != prev_op:
        # operator group header
        grp_row(ws1, row, f"  {op}", NC1, bg=P["blue_acc"])
        row += 1
        prev_op = op
        alt = False

    sn += 1
    used_key = used.split(" ")[0]
    bg_used, fg_used = USED_MAP.get(used_key, (P["white"], P["slate"]))
    row_bg = P["slate_lt"] if alt else P["white"]

    vals = [sn, op, folder, fname, fmt,
            f"{size_kb:,}", records, purpose, used]
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(row=row, column=ci)
        if ci == 9:   # Used column — colour-coded
            wcell(c, val, bold=True, sz=9, fg=fg_used, bg=bg_used)
        elif ci in (1, 2):
            wcell(c, val, bold=True, sz=9, bg=row_bg)
        else:
            wcell(c, val, sz=9, bg=row_bg)
        if ci == 1:
            c.alignment = Alignment(horizontal="center", vertical="top")

    ws1.row_dimensions[row].height = 44
    alt = not alt
    row += 1

# Legend
row += 1
for label, (bg, fg) in [
    ("Yes — fully used on the map", (P["green_lt"], P["green"])),
    ("Partial — only a subset is used (see Sheet 3)", (P["amber_lt"], P["amber"])),
    ("No — data not yet on the map", (P["red_lt"], P["red"])),
    ("Indirect — used as a source for another file", (P["purple_lt"], P["purple"])),
]:
    ws1.merge_cells(f"A{row}:I{row}")
    c = ws1.cell(row=row, column=1)
    wcell(c, f"  ■  {label}", bold=True, sz=9, fg=fg, bg=bg, border=THIN)
    ws1.row_dimensions[row].height = 14
    row += 1

ws1.auto_filter.ref = f"A4:{get_column_letter(NC1)}4"


# ════════════════════════════════════════════════════════════════════
# SHEET 2 — Data Types Found
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Data Types Found")

COLS2 = [
    ("S/N",              4),
    ("Operator",        20),
    ("Dataset",         22),
    ("Geometry",        11),
    ("Total in Source", 13),
    ("Sub-types / Categories",       42),
    ("Key Attributes",               36),
    ("Coord System",     14),
    ("Data Period",      12),
]
NC2 = len(COLS2)
title_row(ws2, 1, "BTRC IMS — Data Types Found per Dataset", NC2)
sub_row(ws2, 2, "What data types and sub-categories exist inside each received file.", NC2)
ws2.row_dimensions[3].height = 5
hdr_row(ws2, 4, COLS2)
set_widths(ws2, COLS2)
ws2.freeze_panes = "A5"

DATA_TYPES = [
    # (operator, dataset, geom, total, subtypes, key_attrs, crs, period)
    ("Grameenphone (GP)", "BTS Sites",
     "Point", "23,864",
     "Fiber Connected: 8,971\nMW Connected: 14,893",
     "Site Code, Latitude, Longitude, TX Type",
     "WGS84", "Dec 2025"),
    ("Robi Axiata", "BTS Sites",
     "Point", "18,838",
     "MW Only: 13,760\nFiber Only: 1,632\nMW + Fiber (Both): 3,446",
     "Site Code, Latitude, Longitude, MW flag (Yes/–), Fiber flag (Yes/–)",
     "WGS84", "Nov 2025"),
    ("Banglalink", "BTS Sites (Latest)",
     "Point", "15,154",
     "Microwave: 13,015\nFiber: 2,113\nInactive: 26",
     "Unique Site ID, Lat, Lon, Division, District, Thana, Address, MW Link Status, Fiber Optic Status, Primary Backhaul Type",
     "WGS84", "Dec 2025"),
    ("Banglalink", "BTS Towers (Historical)",
     "Point", "13,208",
     "4G sites\n3G only\n2G only",
     "Site Code, Site Name, Generation (2G/3G/4G), Division, District, Upazila, Union, Vendor",
     "WGS84", "Pre-2025"),
    ("Banglalink", "Fiber Lines (Historical)",
     "LineString", "172",
     "72 Core\n48 Core\n32 Core",
     "Core Count",
     "WGS84", "Pre-2025"),
    ("BTCL", "Network Points (Latest Excel)",
     "Point", "24,142",
     "CP (Connection Point)\nHH (Hand Hole)\nHOP\nPOP\nMH (Man Hole)\nOther\n(3,432 raw type values normalised to 6 categories)",
     "Name, Latitude, Longitude, Point Type, Raw Type, Year, Feature Code",
     "WGS84", "2025"),
    ("BTCL", "Fiber Lines (Historical GeoJSON)",
     "MultiLineString", "584",
     "144+ Core\n96 Core\n48 Core\n24 Core\n<24 Core",
     "Operator Name, Line Name, Line Type, Core Count, Route Length (km), Cable Length (km), Duct Info, Division/District/Upazila",
     "WGS84", "Pre-2025"),
    ("BTCL", "Network Nodes (Historical GeoJSON)",
     "Point", "29,795",
     "HOP\nHH (Handhole)\nCP (Connection Point)\nMH (Manhole)",
     "Operator Name, Point Name, Point Type, Lat, Lon, Division, District, Upazila, Union, Mouza, Year",
     "WGS84", "Pre-2025"),
    ("BTCL", "Union Project Locations",
     "Point", "966",
     "Union-level project markers",
     "Location coordinates",
     "WGS84", "Pre-2025"),
    ("Fiber@Home (FHLFON)", "Fiber Lines",
     "LineString", "141,567",
     "Burial: 53,870\nAerial: 85,366\nOPGW: 2,331",
     "Operator Code, Operator Name, Infra Type (Own/Leased), Line ID, Year, Month, Line Name, Line Type, Path Along, Duct No, Duct Use, Cable No",
     "WGS84", "Dec 2024"),
    ("Fiber@Home (FHLFON)", "Network Points",
     "Point", "122,150",
     "JE (Joint Enclosure): 38,097\nEP (End Point): 25,804\nCoupler: 22,435\nFL (Fiber Loop): 16,558\nHH (Hand Hole): 8,292\nBTS: 4,755\nCO (Central Office): 2,087\nFAT: 1,591\nBDB: 1,390\nPole: 479\nLDP: 387\nFDH: 80\nPIT: 26\nCustomer: 25\nOther/Noise: ~44",
     "Year, Month, Operator Code, Operator Name, Point ID, Point Type, Point Name, Address, Latitude, Longitude, Feature Code",
     "WGS84", "Dec 2024"),
    ("Summit Communications", "Fiber Lines",
     "LineString", "102,514",
     "Aerial: 78,499\nBurial: 22,702\nOverhead PGCB: 1,209\nBridge Crossing: 14\nBurial (Damaged): 89",
     "FID, Line ID, Year, Month, Operator Code, Operator Name, Infra Type, Line Type, Line Name, Path Along, No of Ducts, Duct Use",
     "WGS84", "Mar 2026"),
    ("Summit Communications", "Network Points",
     "Point", "68,850",
     "TJB (Joint Box): 37,909\nHH (Hand Hole): 9,506\nBTS: 8,613\nEP (End Point): 6,872\nNode/Info: 1,539\nODB: 1,193\nNode/TT: 812\nPOP: 666\nPOC: 453\nBS (Base Station): 422\nNode/CBD: 245\nCOLO/PoP: 128\nFAT: 89\nNode: 77\nNode/GP HO: 76\nFDT: 62\nOther: ~183",
     "FID, Point ID, Year, Month, Operator Code, Operator Name, Point Type, Point Name, Latitude, Longitude, Feature Code",
     "WGS84", "Mar 2026"),
    ("Bahon Limited", "Fiber Lines",
     "LineString", "~7,763\n(from shapefile)",
     "Overhead (OH)\nUnderground (UG)\nWall Clamped (WC)",
     "Cable Type, Division, District, Upazila (from DBF attributes)",
     "WGS84", "Pre-2025"),
    ("Bahon Limited", "Network Nodes",
     "Point", "12,817",
     "Network junction nodes",
     "Node coordinates",
     "WGS84", "Pre-2025"),
    ("InfoSarkar-3 (IS3)", "Fiber Lines",
     "LineString", "3,383",
     "48 Core\n24 Core\n12 Core\nMessenger\nRing\nCBD",
     "Core count, Name, Layer, Length (km)",
     "WGS84", "Pre-2025"),
    ("InfoSarkar-3 (IS3)", "Network Nodes",
     "Point", "477",
     "Network junction nodes",
     "Node coordinates",
     "WGS84", "Pre-2025"),
    ("PGCB", "OPGW Transmission Lines",
     "LineString", "324",
     "400 kV Transmission Line\n230 kV Transmission Line\n132 kV Transmission Line\nUnderground Cable\nOthers",
     "Layer (voltage/type), Name, Description",
     "WGS84", "Pre-2025"),
    ("Multi-operator\n(GP, Robi, BTCL, BL, MOTN, BSCCL)", "Fiber Network Lines",
     "LineString", "8,163",
     "By operator:\nGrameenphone, Robi, BTCL,\nBanglalink, MOTN, BSCCL, Unknown",
     "Name, Operator, Distance (km)",
     "WGS84", "2025"),
    ("Multi-operator\n(GP, Robi, BTCL, BL, MOTN, BSCCL)", "Fiber Network Points",
     "Point", "19,096",
     "By operator:\nGrameenphone, Robi, BTCL,\nBanglalink, MOTN, BSCCL, Unknown",
     "Name, Operator",
     "WGS84", "2025"),
    ("Bangladesh Railway (BR)", "Fiber Lines",
     "LineString", "353",
     "8 Core\n16 Core\n32 Core\n48 Core\n72 Core\n96 Core",
     "Station Name A, Station Name B, Length (km), Total Core, Used Core, Unused Core",
     "WGS84", "2025"),
    ("Bangladesh Railway (BR)", "Station Fiber Nodes",
     "Point", "354",
     "Station junction nodes",
     "Station name",
     "WGS84", "2025"),
    ("Bangladesh Railway (BR)", "Railway Track",
     "LineString", "2,675",
     "Rail route segments",
     "Route geometry",
     "WGS84", "Pre-2025"),
    ("All ISPs (Multi-operator)", "ISP POP Locations",
     "Point", "3,930",
     "Category-A ISPs\nCategory-B ISPs\nCategory-C ISPs\n(covers all districts nationwide)",
     "GID, POP Code, POP Name, POP Address, POP Capacity, ISP Name, Type of ISP, Lat, Lon, District, Division, Upazila, Union, Mouza",
     "WGS84", "2020–2021"),
]

row = 5
alt = False
prev_op = None
for sn, (op, dataset, geom, total, subtypes, attrs, crs, period) in enumerate(DATA_TYPES, 1):
    if op != prev_op:
        grp_row(ws2, row, f"  {op}", NC2, bg=P["blue_acc"])
        row += 1
        prev_op = op
        alt = False

    row_bg = P["slate_lt"] if alt else P["white"]
    vals = [sn, op, dataset, geom, total, subtypes, attrs, crs, period]
    for ci, val in enumerate(vals, 1):
        c = ws2.cell(row=row, column=ci)
        wcell(c, val, bold=(ci <= 2), sz=9, bg=row_bg,
              align="center" if ci == 1 else "left")
    ws2.row_dimensions[row].height = 72
    alt = not alt
    row += 1

ws2.auto_filter.ref = f"A4:{get_column_letter(NC2)}4"
set_widths(ws2, COLS2)
ws2.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════
# SHEET 3 — Implementation Status
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Implementation Status")

COLS3 = [
    ("S/N",              4),
    ("Operator",        20),
    ("Dataset",         22),
    ("Source Records",  13),
    ("On Map",          13),
    ("Difference",      11),
    ("% Coverage",      11),
    ("Map Layer Name",  26),
    ("Status",          13),
    ("Gap / Missing",   44),
]
NC3 = len(COLS3)
title_row(ws3, 1, "BTRC IMS — Map Implementation Status vs. Source Data", NC3)
sub_row(ws3, 2, "Comparison of what is available in received files vs. what is currently rendered on the map.", NC3)
ws3.row_dimensions[3].height = 5
hdr_row(ws3, 4, COLS3)
set_widths(ws3, COLS3)
ws3.freeze_panes = "A5"

STATUS_ROWS = [
    # (op, dataset, source, on_map, diff, pct, layer, status, gap)
    ("Grameenphone (GP)", "BTS Sites",
     "23,864", "23,864", "0", "100%",
     "Grameenphone → Sites",
     "Complete",
     "All sites rendered. Admin (Div/Dist/Upazila) enriched via PIP. 246 border-edge sites show '—' for admin."),
    ("Robi Axiata", "BTS Sites",
     "18,838", "18,838", "0", "100%",
     "Robi → Sites",
     "Complete",
     "All sites rendered. Admin enriched via PIP. 169 border-edge sites show '—'."),
    ("Banglalink", "BTS Sites (Latest)",
     "15,154", "15,154", "0", "100%",
     "BL → BTS Sites (Latest)",
     "Complete",
     "All sites rendered. Admin data taken directly from Excel (no PIP needed)."),
    ("Banglalink", "BTS Towers (Historical)",
     "13,208", "13,208", "0", "100%",
     "BL → Towers",
     "Complete",
     "All tower sites with 2G/3G/4G generation filter."),
    ("Banglalink", "Fiber Lines (Historical)",
     "172", "172", "0", "100%",
     "BL → Fiber Lines",
     "Complete",
     "All lines with 32/48/72 core filter."),
    ("BTCL", "Network Points (Latest Excel)",
     "24,142", "24,073", "69", "99.7%",
     "BTCL → Points (Excel 2025)",
     "Complete*",
     "69 points removed by PIP filter (coordinates outside Bangladesh polygon). Acceptable data quality loss."),
    ("BTCL", "Fiber Lines (Historical)",
     "584", "584", "0", "100%",
     "BTCL-OLD → Fiber Lines",
     "Complete",
     "All lines rendered with core count filter."),
    ("BTCL", "Network Nodes (Historical)",
     "29,795", "29,795", "0", "100%",
     "BTCL-OLD → Nodes",
     "Complete",
     "All nodes with HOP/HH/CP/MH type filter."),
    ("BTCL", "Union Project Locations",
     "966", "966", "0", "100%",
     "BTCL-OLD → Union Projects",
     "Complete",
     "All union project markers rendered."),
    ("Fiber@Home (FHLFON)", "Fiber Lines",
     "141,567", "18,405", "123,162", "13%",
     "Fiber@Home → Lines",
     "Partial",
     "Only a pre-converted subset is on the map. Full shapefile (141,567 lines incl. OPGW: 2,331) not yet re-ingested from source. Missing: Aerial lines beyond current subset, all OPGW type lines."),
    ("Fiber@Home (FHLFON)", "Network Points",
     "122,150", "94,953", "27,197", "78%",
     "Fiber@Home → Points",
     "Partial",
     "Missing point types: Coupler (22,435), FL (16,558), BDB (1,390), Pole (479), LDP (387), PIT (26), Customer (25). Currently map shows CO/BTS/FDH/JE/EP/FAT only."),
    ("Summit Communications", "Fiber Lines",
     "102,514", "23,157", "79,357", "23%",
     "Summit → Lines",
     "Partial",
     "Only a pre-classified subset (Backbone/Major/PGCB/Railway tiers) is on the map. Full Excel has 102,514 lines across Aerial (78,499), Burial (22,702), Overhead_PGCB (1,209), Bridge Crossing (14), Burial(Damaged) (89). Missing ~79k lines."),
    ("Summit Communications", "Network Points",
     "68,850", "14,562", "54,288", "21%",
     "Summit → Nodes + BTS",
     "Partial",
     "Only Nodes (5,949) and BTS (8,613) on map. Missing: TJB (37,909), EP (6,872), ODB (1,193), Node/TT (812), POP (666), POC (453), BS (422), Node/CBD (245), COLO/PoP (128), FAT (89), FDT (62) and others."),
    ("Bahon Limited", "Fiber Lines",
     "~7,763\n(shapefile)", "7,763", "~0", "~100%",
     "Bahon → Lines",
     "Complete*",
     "Appears complete. Note: exact source record count not verified from DBF (shapefile only, no Excel counterpart)."),
    ("Bahon Limited", "Network Nodes",
     "12,817", "12,817", "0", "100%",
     "Bahon → Nodes",
     "Complete",
     "All nodes rendered."),
    ("InfoSarkar-3 (IS3)", "Fiber Lines",
     "3,383", "3,383", "0", "100%",
     "InfoSarkar-3 → Lines",
     "Complete",
     "All lines with core count filter (48/24/12/Messenger/Ring/CBD)."),
    ("InfoSarkar-3 (IS3)", "Network Nodes",
     "477", "477", "0", "100%",
     "InfoSarkar-3 → Nodes",
     "Complete",
     "All nodes rendered."),
    ("PGCB", "OPGW Lines",
     "324", "324", "0", "100%",
     "PGCB → Lines",
     "Complete",
     "All OPGW lines with voltage-level filter."),
    ("Multi-operator", "Fiber Lines",
     "8,163", "8,163", "0", "100%",
     "Fiber Network → Lines",
     "Complete",
     "All 6-operator fiber lines with operator colour filter."),
    ("Multi-operator", "Fiber Points",
     "19,096", "19,096", "0", "100%",
     "Fiber Network → Points",
     "Complete",
     "All 6-operator fiber junction points."),
    ("Bangladesh Railway (BR)", "Fiber Lines",
     "353", "353", "0", "100%",
     "Railway → BR Fiber Lines",
     "Complete",
     "All lines with core count filter."),
    ("Bangladesh Railway (BR)", "Station Nodes",
     "354", "354", "0", "100%",
     "Railway → BR Fiber Nodes",
     "Complete",
     "All station nodes rendered."),
    ("Bangladesh Railway (BR)", "Railway Track",
     "2,675", "2,675", "0", "100%",
     "Railway → Railline",
     "Complete",
     "Base railway track layer fully rendered."),
    ("All ISPs", "ISP POP Locations",
     "3,930", "0", "3,930", "0%",
     "— Not implemented —",
     "Not Started",
     "Full data available in aisp-pop.geojson with ISP name, capacity, category, and full admin data. Not yet added to map."),
]

STATUS_COLORS = {
    "Complete":     (P["green_lt"],  P["green"]),
    "Complete*":    (P["green_lt"],  P["green"]),
    "Partial":      (P["amber_lt"],  P["amber"]),
    "Not Started":  (P["red_lt"],    P["red"]),
}

row = 5
alt = False
prev_op = None
for sn, (op, dataset, source, on_map, diff, pct, layer, status, gap) in enumerate(STATUS_ROWS, 1):
    if op != prev_op:
        grp_row(ws3, row, f"  {op}", NC3, bg=P["blue_acc"])
        row += 1
        prev_op = op
        alt = False

    bg_s, fg_s = STATUS_COLORS.get(status, (P["white"], P["slate"]))
    row_bg = P["slate_lt"] if alt else P["white"]

    vals = [sn, op, dataset, source, on_map, diff, pct, layer, status, gap]
    for ci, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=ci)
        if ci == 9:
            wcell(c, val, bold=True, sz=9, fg=fg_s, bg=bg_s, align="center")
        elif ci == 6 and val not in ("0", "~0"):
            wcell(c, val, bold=True, sz=9, fg=P["red"], bg=row_bg)
        elif ci == 7 and val not in ("100%", "~100%", "99.7%"):
            wcell(c, val, bold=True, sz=9, fg=P["amber"], bg=row_bg)
        else:
            wcell(c, val, bold=(ci <= 2), sz=9, bg=row_bg,
                  align="center" if ci == 1 else "left")

    ws3.row_dimensions[row].height = 52
    alt = not alt
    row += 1

ws3.auto_filter.ref = f"A4:{get_column_letter(NC3)}4"
set_widths(ws3, COLS3)
ws3.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════
# SHEET 4 — Missing for Complete Map
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Missing for Complete Map")

COLS4 = [
    ("Priority",         10),
    ("Category",         18),
    ("Operator",         18),
    ("Missing Item",     36),
    ("Data Available?",  15),
    ("Est. Records",     14),
    ("Effort",           10),
    ("Action Required",  46),
]
NC4 = len(COLS4)
title_row(ws4, 1, "BTRC IMS — Missing Items for Complete Map Implementation", NC4)
sub_row(ws4, 2, "Prioritised list of what remains to be done for a fully complete map.", NC4)
ws4.row_dimensions[3].height = 5
hdr_row(ws4, 4, COLS4)
set_widths(ws4, COLS4)
ws4.freeze_panes = "A5"

PRI_COLORS = {
    "High":   (P["red_lt"],    P["red"]),
    "Medium": (P["amber_lt"],  P["amber"]),
    "Low":    (P["green_lt"],  P["green"]),
    "Info":   (P["purple_lt"], P["purple"]),
}
AVAIL_COLORS = {
    "Yes":            P["green"],
    "Yes (partial)":  P["amber"],
    "No":             P["red"],
}

MISSING = [
    # (priority, category, operator, missing_item, available, est_records, effort, action)
    ("High", "New Layer", "All ISPs",
     "ISP POP Locations not on map",
     "Yes", "3,930",
     "Low",
     "Add ISP POP overlay using aisp-pop.geojson. Filter by ISP category (A/B/C) and district. Popup should show ISP name, POP name, address, capacity, admin data."),

    ("High", "Incomplete Data", "Fiber@Home (FHLFON)",
     "Fiber Lines: 123,162 lines missing (only 13% on map)",
     "Yes", "141,567 total\n(123,162 missing)",
     "High",
     "Re-ingest full FHLFONLineExcel.xlsx or shapefile. Current map has only a pre-converted subset. Need to convert all 141,567 lines including OPGW type (2,331)."),

    ("High", "Incomplete Data", "Summit Communications",
     "Fiber Lines: 79,357 lines missing (only 23% on map)",
     "Yes", "102,514 total\n(79,357 missing)",
     "High",
     "Re-ingest full Line_data.xlsx or shapefile. Current map uses a simplified 3-tier classification. Full data has Aerial (78k), Burial (22k), Overhead_PGCB, Bridge Crossing types."),

    ("High", "Incomplete Data", "Summit Communications",
     "Network Points: 54,288 points missing (only 21% on map)",
     "Yes", "68,850 total\n(54,288 missing)",
     "High",
     "Re-ingest full Point_data.xlsx. Missing types: TJB (37,909), EP (6,872), ODB (1,193), Node/TT (812), POP (666), POC (453), BS (422), Node/CBD (245), COLO/PoP (128), FAT (89), FDT (62)."),

    ("Medium", "Incomplete Data", "Fiber@Home (FHLFON)",
     "Network Points: 7 point types not shown (Coupler, FL, BDB, Pole, LDP, PIT, Customer)",
     "Yes", "~41,407 missing\npoints",
     "Medium",
     "Add missing point types to FHLFON overlay filter. Source data in FHLFONPointExcel.xlsx. Current map shows only CO/BTS/FDH/JE/EP/FAT. Missing: Coupler (22,435), FL (16,558), BDB (1,390), Pole (479), LDP (387), PIT (26), Customer (25)."),

    ("Medium", "Missing Archive", "Summit Communications",
     "Summit.rar in Railway folder not extracted or analysed",
     "Yes (archive)", "Unknown",
     "Low",
     "Extract and analyse Summit.rar to determine if it contains additional network data not present in the Excel/shapefile files."),

    ("Medium", "Missing Submission", "Bahon Limited",
     "No Excel tabular data received — shapefile only",
     "Yes (partial)", "~7,763 lines",
     "Low",
     "Bahon submitted only shapefiles with no Excel counterpart. Request Excel tabular submission from Bahon for BTRC records. Shapefile is implemented but exact record count unverified."),

    ("Low", "Attribute Gap", "Grameenphone (GP)",
     "No 2G/3G/4G technology generation in BTS submission",
     "No", "—",
     "Low",
     "GP Dec-2025 submission only includes TX type (Fiber/MW). Generation info (2G/3G/4G) not provided. Request updated submission with technology generation field."),

    ("Low", "Attribute Gap", "Robi Axiata",
     "No technology generation (2G/3G/4G) in BTS submission",
     "No", "—",
     "Low",
     "Robi Nov-2025 submission only has MW/Fiber backhaul flags. No generation info. Request updated submission."),

    ("Low", "Attribute Gap", "Banglalink",
     "Latest BTS submission has no vendor or technology generation info",
     "No", "—",
     "Low",
     "Latest BTS submission (Dec 2025) lacks Vendor and Generation fields present in the historical tower dataset. Request updated fields in future submissions."),

    ("Low", "Missing Operator", "Teletalk",
     "No data received from Teletalk",
     "No", "~5,000 est.",
     "—",
     "No BTS, fiber, or network data received from Teletalk (state-owned mobile operator). BTRC should request submission."),

    ("Low", "Missing Operator", "Airtel / Other MVNOs",
     "No data from virtual/small operators",
     "No", "Unknown",
     "—",
     "No data from MVNOs or smaller ISPs beyond the POP dataset. Confirm scope with BTRC."),

    ("Info", "Data Quality", "BTCL",
     "69 BTCL points outside Bangladesh boundary removed",
     "Yes (filtered)", "69 removed",
     "—",
     "69 points from the latest Excel were filtered out by point-in-polygon check. Coordinates appear to be data entry errors near borders. BTCL should verify and resubmit corrected coordinates."),

    ("Info", "Data Quality", "Fiber@Home (FHLFON)",
     "Copy of FHLFONPointExcel.xlsx is an exact duplicate",
     "N/A", "—",
     "—",
     "File 'Copy of FHLFONPointExcel.xlsx' is identical to 'FHLFONPointExcel.xlsx'. No additional data. Can be ignored."),

    ("Info", "Data Quality", "Bangladesh Railway",
     "Railway Excel duplicated in two folders",
     "N/A", "—",
     "—",
     "Geo Spatial Data Structure_Template_Final_railway.xlsx appears in both 'Latest Data/' and 'Railway/' with same content. Only one copy used."),
]

row = 5
for sn, (priority, category, operator, item, available, est, effort, action) in enumerate(MISSING, 1):
    bg_p, fg_p = PRI_COLORS.get(priority, (P["white"], P["slate"]))
    fg_a = AVAIL_COLORS.get(available.split(" ")[0], P["slate"])
    row_bg = P["slate_lt"] if sn % 2 == 0 else P["white"]

    vals = [priority, category, operator, item, available, est, effort, action]
    for ci, val in enumerate(vals, 1):
        c = ws4.cell(row=row, column=ci)
        if ci == 1:
            wcell(c, val, bold=True, sz=9, fg=fg_p, bg=bg_p, align="center")
        elif ci == 5:
            wcell(c, val, bold=True, sz=9, fg=fg_a, bg=row_bg, align="center")
        elif ci == 7:
            effort_fg = P["red"] if val == "High" else P["amber"] if val == "Medium" else P["green"] if val == "Low" else P["slate"]
            wcell(c, val, bold=True, sz=9, fg=effort_fg, bg=row_bg, align="center")
        else:
            wcell(c, val, bold=(ci == 4), sz=9, bg=row_bg,
                  align="center" if ci in (1,) else "left")
    ws3.row_dimensions[row].height = 60
    ws4.row_dimensions[row].height = 60
    row += 1

set_widths(ws4, COLS4)
ws4.freeze_panes = "A5"
ws4.auto_filter.ref = f"A4:{get_column_letter(NC4)}4"

# ── Page setup all sheets ─────────────────────────────────────────
for ws in [ws1, ws2, ws3, ws4]:
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.print_title_rows        = "1:4"
    ws.sheet_view.showGridLines = True

wb.save(str(OUT))
print(f"Saved -> {OUT}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
